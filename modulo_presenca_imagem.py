import json
import os
import datetime
import cv2
import numpy as np
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.dropdown import DropDown
from kivy.uix.image import Image
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from openpyxl import Workbook

import face_recognition

from helper_funcoes_reutilizadas import helper_busca_disciplinas


class CapturaPresencaImagem(App):
    def build(self):
        self.title = 'Instituto Federal Fluminense - Gerar Presença por Imagem'
        self.icon = 'Imagens/icone_camera.png'
        self.image = Image(source='Imagens/selecione_imagem.png')

        try:
            with open("Configuracoes/config.txt", "r") as config_file:
                config = json.load(config_file)
                select_matricula = config.get("select_matricula", "")
                select_disciplina = config.get("select_disciplina", "")
                select_curso = config.get("select_curso", "")
        except FileNotFoundError:
            select_disciplina = ""
            select_curso = ""

        left_column = BoxLayout(orientation='vertical', spacing=10)
        left_column.add_widget(self.image)
        self.file_chooser = FileChooserListView(path=str("CapturasDeTurma"))
        self.file_chooser.bind(on_submit=self.load_image)
        left_column.add_widget(self.file_chooser)

        self.label_date = Label(text='Data:')
        self.label_text_input1 = Label(text='Curso:')
        self.label_text_input2 = Label(text='Disciplina:')
        data_de_hoje = datetime.datetime.now().strftime('%Y-%m-%d')
        self.date_input = TextInput(hint_text='Data', text=data_de_hoje)
        self.text_input1 = TextInput(hint_text='Curso', text=select_curso)

        self.text_input2 = TextInput(size_hint_y=None, height='48dp', hint_text='Selecione a Disciplina',
                                     readonly=True)

        helper = helper_busca_disciplinas()
        lista_disciplinas = helper.lista_de_disciplinas_cadastradas()
        dropdown = DropDown()

        for disciplina in lista_disciplinas:
            btn = Button(text=disciplina, size_hint_y=None, height=44)
            btn.bind(on_release=lambda btn: (setattr(self, 'select_disciplina', btn.text),
                                             setattr(self.text_input2, 'text', btn.text),
                                             dropdown.dismiss()))
            dropdown.add_widget(btn)

        self.text_input2.bind(
            on_touch_down=lambda instance, touch: dropdown.open(self.text_input2) if instance.collide_point(
                *touch.pos) else None)

        generate_button = Button(text='Gerar Presença')
        close_button = Button(text='Fechar')
        generate_button.bind(on_press=self.generate_presence)
        close_button.bind(on_press=self.close_popup)

        right_column = BoxLayout(orientation='vertical', spacing=10)
        right_column.add_widget(self.label_date)
        right_column.add_widget(self.date_input)
        right_column.add_widget(self.label_text_input1)
        right_column.add_widget(self.text_input1)
        right_column.add_widget(self.label_text_input2)
        right_column.add_widget(self.text_input2)
        right_column.add_widget(generate_button)
        right_column.add_widget(close_button)

        layout = BoxLayout(spacing=10)
        layout.add_widget(left_column)
        layout.add_widget(right_column)

        return layout

    def load_image(self, instance, selection, *args):
        if selection:
            value = selection[0]
            if os.path.isfile(value) and value.lower().endswith(('.png', '.jpg', '.jpeg')):
                self.image.source = value
            else:
                self.show_error_popup("Selecione um arquivo de imagem válido.")

    def generate_presence(self, instance):
        data = self.date_input.text
        curso = self.text_input1.text
        disciplina = self.text_input2.text

        if not disciplina:
            self.show_error_popup("Por favor, selecione uma disciplina antes de gerar a presença.")
            return

        if "selecione_imagem.png" in self.image.source:
            self.show_error_popup("Por favor, selecione uma imagem antes de gerar a presença.")
            return

        people_dir = os.path.join("Alunos", disciplina)

        known_faces, known_names = self.load_known_faces(people_dir)

        if not known_faces:
            self.show_error_popup("Nenhum rosto conhecido foi carregado. Verifique as imagens dos alunos.")
            return

        all_people = [os.path.splitext(filename)[0] for filename in os.listdir(people_dir)
                      if filename.endswith(('.png', '.jpg', '.jpeg'))]

        recognized_people = self.get_pessoas_presentes(self.image.source, known_faces, known_names)

        filename = f"{curso}_{disciplina}_{data}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value="Aluno")
        sheet.cell(row=1, column=2, value="Matricula")
        sheet.cell(row=1, column=3, value="Status")

        row_index = 2

        for person in all_people:
            person = person.replace(".png", "").replace(".jpg", "").replace(".jpeg", "")

            try:
                _, nome, matricula = person.split("_", 2)
            except ValueError:
                print(f"Formato inválido para o arquivo: {person}")
                continue

            status = "PRESENTE" if person in recognized_people else "AUSENTE"

            sheet.cell(row=row_index, column=1, value=nome)
            sheet.cell(row=row_index, column=2, value=matricula)
            sheet.cell(row=row_index, column=3, value=status)
            row_index += 1

        app_root_dir = os.path.dirname(os.path.abspath(__file__))
        save_dir = os.path.join(app_root_dir, "PresencasCapturadas")
        os.makedirs(save_dir, exist_ok=True)

        file_path = os.path.join(save_dir, filename)
        workbook.save(file_path)

        self.show_info_popup(f'Arquivo de presença "{filename}" \nfoi gerado com sucesso na pasta.')

    def load_known_faces(self, people_dir):
        known_faces = []
        known_names = []

        if not os.path.exists(people_dir):
            self.show_error_popup(f"A pasta '{people_dir}' não existe.")
            return [], []

        for person_filename in os.listdir(people_dir):
            person_path = os.path.join(people_dir, person_filename)
            if person_filename.endswith(('.png', '.jpg', '.jpeg')):
                image = face_recognition.load_image_file(person_path)
                encodings = face_recognition.face_encodings(image)

                if len(encodings) > 0:
                    known_faces.append(encodings[0])
                    person_name = os.path.splitext(person_filename)[0]
                    known_names.append(person_name)
                else:
                    print(f"Nenhum rosto detectado em: {person_filename}")

        return known_faces, known_names

    def get_pessoas_presentes(self, image_path, known_faces, known_names):
        image = face_recognition.load_image_file(image_path)
        face_locations = face_recognition.face_locations(image)
        face_encodings = face_recognition.face_encodings(image, face_locations)

        recognized_people = []

        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(known_faces, face_encoding)
            if any(matches):
                face_index = matches.index(True)
                person_name = known_names[face_index]
                recognized_people.append(person_name)

        return recognized_people

    def show_error_popup(self, message):
        content = BoxLayout(orientation='vertical')
        content.add_widget(Label(text=message))
        close_button = Button(text='Fechar')
        content.add_widget(close_button)
        popup = Popup(title='Erro', content=content, auto_dismiss=False, size_hint=(None, None), size=(400, 300))
        close_button.bind(on_release=popup.dismiss)
        popup.open()

    def show_info_popup(self, message):
        content = BoxLayout(orientation='vertical')
        content.add_widget(Label(text=message))
        close_button = Button(text='Fechar')
        content.add_widget(close_button)
        popup = Popup(title='Informação', content=content, auto_dismiss=False, size_hint=(None, None), size=(600, 300))
        close_button.bind(on_release=popup.dismiss)
        popup.open()

    def close_popup(self, instance):
        App.get_running_app().stop()


if __name__ == '__main__':
    CapturaPresencaImagem().run()
