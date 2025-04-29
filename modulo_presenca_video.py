import json
import cv2
import face_recognition
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.dropdown import DropDown
from kivy.uix.popup import Popup
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.image import Image
from kivy.uix.label import Label
from kivy.clock import Clock
from kivy.graphics.texture import Texture
from datetime import datetime
from openpyxl import Workbook
import os
from openpyxl.reader.excel import load_workbook
from helper_funcoes_reutilizadas import helper_busca_disciplinas


class CapturaPresencaVideo(App):
    def build(self):
        self.title = 'Instituto Federal Fluminense - Captura de Presença'
        self.icon = 'Imagens/icone_camera.png'

        try:
            with open("Configuracoes/config.txt", "r") as config_file:
                config = json.load(config_file)
                select_cam = config.get("select_cam", "")
                select_matricula = config.get("select_matricula", "")
                select_disciplina = config.get("select_disciplina", "")
                select_curso = config.get("select_curso", "")
        except FileNotFoundError:
            pass

        layout = BoxLayout(orientation='horizontal', spacing=10)

        col1_layout = BoxLayout(orientation='vertical', spacing=10)

        self.camera = Image(source='Imagens/no_camera.png', allow_stretch=True)
        col1_layout.add_widget(self.camera)

        self.timer_label = Label(text="Tempo restante: 10:00")
        col1_layout.add_widget(self.timer_label)

        layout.add_widget(col1_layout)

        col2_layout = BoxLayout(orientation='vertical', spacing=10)

        label_date = Label(text='Data:')
        label_text_input1 = Label(text='Curso:')
        label_text_input2 = Label(text='Disciplina:')

        data_de_hoje = datetime.now().strftime('%Y-%m-%d')
        date_input = TextInput(hint_text='Data', text=data_de_hoje)

        text_input1 = TextInput(hint_text='Curso', text=select_curso)

        self.disciplina_input = TextInput(size_hint_y=None, height='48dp', hint_text='Selecione a Disciplina',
                                          readonly=True)
        helper = helper_busca_disciplinas()
        lista_disciplinas = helper.lista_de_disciplinas_cadastradas()
        dropdown = DropDown()

        for disciplina in lista_disciplinas:
            btn = Button(text=disciplina, size_hint_y=None, height=44)
            btn.bind(on_release=lambda btn: (setattr(self, 'select_disciplina', btn.text),
                                             setattr(self.disciplina_input, 'text', btn.text),
                                             dropdown.dismiss()))
            dropdown.add_widget(btn)

        self.disciplina_input.bind(
            on_touch_down=lambda instance, touch: dropdown.open(self.disciplina_input) if instance.collide_point(
                *touch.pos) else None)

        col2_layout.add_widget(label_date)
        col2_layout.add_widget(date_input)
        col2_layout.add_widget(label_text_input1)
        col2_layout.add_widget(text_input1)
        col2_layout.add_widget(label_text_input2)
        col2_layout.add_widget(self.disciplina_input)

        self.timer_input = TextInput(hint_text='Tempo (segundos)', input_filter='int')
        self.timer_input.text = '600'
        label_timer_input = Label(text='Defina o tempo de captura em segundos:')
        col2_layout.add_widget(label_timer_input)
        col2_layout.add_widget(self.timer_input)

        start_button = Button(text='Iniciar Captura de Presença')
        start_button.bind(on_press=self.inicia_processo_presenca)

        close_button = Button(text='Fechar')
        close_button.bind(on_press=self.close_app)

        col2_layout.add_widget(start_button)
        col2_layout.add_widget(close_button)

        layout.add_widget(col2_layout)

        self.capture = cv2.VideoCapture(0)
        self.known_faces = []
        self.known_names = []
        self.detected_people = {}
        self.timer = None
        self.remaining_time = 600
        self.mostrar_popup = True
        self.is_capturing = False

        Clock.schedule_interval(self.update, 0 / 30.0)

        self.date_input = date_input
        self.text_input1 = text_input1

        return layout

    def inicia_processo_presenca(self, instance):
        self.lista_presenca = []
        user_time_input = int(self.timer_input.text)
        if user_time_input <= 0:
            print("O tempo deve ser um valor inteiro maior que zero")
            return

        if not self.capture.isOpened():
            print("Erro ao abrir a câmera.")
            return

        if not self.is_capturing:
            self.is_capturing = True
            self.mostrar_popup = True

            self.remaining_time = user_time_input
            self.timer_label.text = f"Tempo restante: {user_time_input // 60:02}:{user_time_input % 60:02}"

            if self.timer is None:
                self.timer = Clock.schedule_interval(self.update_timer, 1)

            data = self.date_input.text
            curso = self.text_input1.text
            disciplina = self.select_disciplina
            filename = f"{curso}_{disciplina}_{data}.xlsx"
            file_path = os.path.join("PresencasCapturadas", filename)

            people_dir = os.path.join("Alunos", disciplina)
            self.known_faces, self.known_names = self.load_known_faces(people_dir)

            if not os.path.exists(file_path):
                self.cria_excel(None)
            else:
                print("Captura de presença iniciada, a captura será finalizada ao fim do temporizador")

    def stop_capturing(self):
        self.is_capturing = False
        self.capture.release()
        self.camera.source = 'Imagens/no_camera.png'

    def update(self, dt):
        if self.is_capturing:
            ret, frame = self.capture.read()
            if ret:
                face_locations = face_recognition.face_locations(frame)
                face_encodings = face_recognition.face_encodings(frame, face_locations)
                newly_detected_people = set()

                for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                    matches = face_recognition.compare_faces(self.known_faces, face_encoding)
                    if any(matches):
                        face_index = matches.index(True)
                        person_name = self.known_names[face_index]
                        newly_detected_people.add(person_name)
                        if person_name not in self.detected_people:
                            print(f"Face reconhecida na câmera: {person_name}")
                            self.montagem_presenca(person_name)
                        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
                    else:
                        print("Face não reconhecida detectada na câmera.")
                        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

                exited_people = set(self.detected_people.keys()) - newly_detected_people
                for person_name in exited_people:
                    print(f"{person_name} saiu do alcance da câmera")

                self.detected_people = {person_name: (top, right, bottom, left) for
                                        (top, right, bottom, left), person_name
                                        in zip(face_locations, newly_detected_people)}

                buffer = cv2.flip(frame, 0).tobytes()
                texture = Texture.create(size=(frame.shape[1], frame.shape[0]), colorfmt='bgr')
                texture.blit_buffer(buffer, colorfmt='bgr', bufferfmt='ubyte')
                self.camera.texture = texture
            else:
                self.camera.source = 'Imagens/no_camera.png'

    def montagem_presenca(self, nome_matricula):
        nome, _ = nome_matricula.split("_", 1)
        if nome not in self.lista_presenca:
            self.lista_presenca.append(nome_matricula)
            print(f"{nome_matricula} adicionado na lista de presença")
        else:
            print(f"{nome_matricula} já está na lista de presença")

    def update_timer(self, dt):
        if self.remaining_time > 0:
            self.remaining_time -= 1
            minutes = self.remaining_time // 60
            seconds = self.remaining_time % 60
            self.timer_label.text = f"Tempo restante: {minutes:02}:{seconds:02}"
        elif self.mostrar_popup is True:
            self.timer_label.text = "Tempo esgotado, presença finalizada"
            self.popup_finaliza_tempo()
            self.stop_capturing()
            self.atualiza_excel()

    def popup_finaliza_tempo(self):
        self.mostrar_popup = False
        message = "Tempo esgotado, presença finalizada"

        box_popup = BoxLayout(orientation='vertical')
        box_popup.add_widget(Label(text=message))
        close_button = Button(text='Fechar')
        box_popup.add_widget(close_button)

        popup = Popup(title='Tempo de Presença Esgotado', content=box_popup, size_hint=(0.5, 0.5))

        close_button.bind(on_release=popup.dismiss)
        popup.open()

    def atualiza_excel(self):
        print("Função de atualização do Excel chamada")

        data = self.date_input.text
        curso = self.text_input1.text
        disciplina = self.select_disciplina
        filename = f"{curso}_{disciplina}_{data}.xlsx"
        file_path = os.path.join("PresencasCapturadas", filename)

        workbook = load_workbook(file_path)
        sheet = workbook.active

        for nome_matricula in self.lista_presenca:
            for i, known_name in enumerate(self.known_names):
                if known_name == nome_matricula:
                    row_index = i + 2
                    sheet.cell(row=row_index, column=3, value="PRESENTE")
                    break

        workbook.save(file_path)


    def cria_excel(self, instance):
        data = self.date_input.text
        curso = self.text_input1.text
        disciplina = self.select_disciplina
        filename = f"{curso}_{disciplina}_{data}.xlsx"

        workbook = Workbook()
        sheet = workbook.active

        sheet.cell(row=1, column=1, value="Aluno")
        sheet.cell(row=1, column=2, value="Matricula")
        sheet.cell(row=1, column=3, value="Status")

        people_dir = os.path.join("Alunos", disciplina)

        self.known_faces, self.known_names = self.load_known_faces(people_dir)

        all_people = [os.path.splitext(filename)[0] for filename in os.listdir(people_dir) if filename.endswith(".png")]
        row_index = 2

        for person in all_people:
            nome_matricula = person.split("_")
            if len(nome_matricula) >= 3:
                nome = nome_matricula[1]
                matricula = nome_matricula[2]
                status = "AUSENTE"

                sheet.cell(row=row_index, column=1, value=nome)
                sheet.cell(row=row_index, column=2, value=matricula)
                sheet.cell(row=row_index, column=3, value=status)
                row_index += 1
            else:
                print(
                    f"Erro no formato do arquivo de imagem: {person}. O formato esperado é 'Disciplina_Nome_Matricula.png'.")

        app_root_dir = os.path.dirname(os.path.abspath(__file__))
        save_dir = os.path.join(app_root_dir, "PresencasCapturadas")
        os.makedirs(save_dir, exist_ok=True)

        file_path = os.path.join(save_dir, filename)
        workbook.save(file_path)

    def load_known_faces(self, people_dir):
        known_faces = []
        known_names = []

        if os.path.exists(people_dir):
            for filename in os.listdir(people_dir):
                if filename.endswith(".png"):
                    path = os.path.join(people_dir, filename)
                    image = face_recognition.load_image_file(path)
                    encodings = face_recognition.face_encodings(image)
                    if encodings:
                        encoding = encodings[0]
                        known_faces.append(encoding)
                        known_names.append(os.path.splitext(filename)[0])
                    else:
                        print(f"Nenhuma face encontrada em {filename}")
        else:
            print(f"Pasta de disciplina não encontrada: {people_dir}")

        return known_faces, known_names

    def close_app(self, instance):
        print("Clicou Fechar App")
        App.get_running_app().stop()


if __name__ == '__main__':
    CapturaPresencaVideo().run()
